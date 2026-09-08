"""campaigns_tabular.py : fetching the states that publish a table.

California needed a bespoke PDF parser. Most states do not: they publish a CSV,
a spreadsheet, or a Socrata dataset, and the whole job is "get rows, map
columns, hand them to from_rows()". This module is that half, written once.

WHY SOCRATA IS WORTH ITS OWN BACKEND. It is not a Pennsylvania quirk -- New
York, Connecticut, Maryland and Washington run the same platform, so the states
after PA cost a resource id rather than a new fetcher. The API is stable and
documented, which is the opposite of the situation with a state that hand-rolls
a download page, and it pages predictably.

THE PART THAT CANNOT BE VERIFIED FROM HERE is never the transport -- it is
WHICH dataset. A Socrata resource id or a download URL is a configuration
value, and guessing one produces an adapter that 404s on first run or, worse,
quietly reads last cycle's file. So this module ships `discover_socrata()`,
which searches a portal's catalogue by keyword and hands back the ids with
their names and row counts. Finding the right dataset is one command rather
than an afternoon, and the adapters that use it say plainly which ids are
confirmed and which are placeholders.

Everything here takes an injectable `fetcher`, so every function is testable
without a network, and no state adapter is ever coupled to how bytes arrive.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any, Callable, Optional

USER_AGENT = "surplus-campaigns/1.0 (+https://event.surpluslayer.com)"
TIMEOUT_S = 30.0

# Socrata caps a single page at 50,000 rows. Ask for that and page until a
# short page comes back.
SOCRATA_PAGE = 50_000
MAX_PAGES = 40

Fetcher = Callable[[str], bytes]


class SourceError(RuntimeError):
    """The upstream did not give us a table we can read."""


class NotConfigured(SourceError):
    """The adapter is written but nobody has told it which dataset to read.

    Its own error, and not just a 404, because those are different problems
    with different fixes: a 404 means the dataset moved, this means the id was
    never filled in. An adapter carrying a placeholder id must say so in one
    line naming the command that finds the real one, rather than failing
    somewhere in the transport where it reads as an outage.
    """


def _default_fetcher(url: str) -> bytes:
    import httpx
    with httpx.Client(timeout=TIMEOUT_S, follow_redirects=True) as client:
        resp = client.get(url, headers={"user-agent": USER_AGENT})
    if resp.status_code >= 400:
        raise SourceError(f"HTTP {resp.status_code} from {url}: {resp.text[:160]}")
    return resp.content


def _fetch(url: str, fetcher: Optional[Fetcher]) -> bytes:
    return (fetcher or _default_fetcher)(url)


# ---------------------------------------------------------------------------
# Socrata
# ---------------------------------------------------------------------------

def socrata_url(domain: str, resource_id: str, *, limit: int = SOCRATA_PAGE,
                offset: int = 0, where: str = "") -> str:
    """Build a Socrata resource URL. Kept separate so tests can assert the
    query without a network, and so a caller can print it for a human."""
    url = (f"https://{domain}/resource/{resource_id}.json"
           f"?$limit={limit}&$offset={offset}")
    if where:
        from urllib.parse import quote
        url = f"{url}&$where={quote(where)}"
    return url


def fetch_socrata(domain: str, resource_id: str, *,
                  where: str = "", fetcher: Optional[Fetcher] = None,
                  max_pages: int = MAX_PAGES) -> list[dict]:
    """Every row of a Socrata dataset, paged.

    Stops on a short page. `max_pages` is a guard against a portal that ignores
    $offset and hands back page one forever -- without it that is an infinite
    loop rather than an error, which is the worse of the two failures.
    """
    rows: list[dict] = []
    for page in range(max_pages):
        url = socrata_url(domain, resource_id,
                          offset=page * SOCRATA_PAGE, where=where)
        raw = _fetch(url, fetcher)
        try:
            batch = json.loads(raw)
        except ValueError as exc:
            raise SourceError(
                f"{domain}/{resource_id} did not return JSON: {exc}") from exc
        if isinstance(batch, dict) and batch.get("error"):
            raise SourceError(
                f"{domain}/{resource_id}: {batch.get('message') or batch}")
        if not isinstance(batch, list):
            raise SourceError(
                f"{domain}/{resource_id} returned {type(batch).__name__}, "
                f"not a list of rows")
        rows.extend(row for row in batch if isinstance(row, dict))
        if len(batch) < SOCRATA_PAGE:
            return rows
    raise SourceError(
        f"{domain}/{resource_id} still returning full pages after "
        f"{max_pages}: refusing to page forever")


def discover_socrata(domain: str, query: str = "candidate", *,
                     fetcher: Optional[Fetcher] = None,
                     limit: int = 20) -> list[dict]:
    """Search a Socrata portal's catalogue for datasets matching `query`.

    This is how you find a resource id without opening a browser. Returns
    {id, name, updated, rows, link} per hit, newest-looking first as the
    catalogue orders them.

    An empty `domain` is REFUSED rather than sent. Socrata treats "domains="
    as no filter at all and happily searches every portal in the country, so
    the request succeeds and returns candidate datasets belonging to other
    states -- results that look entirely correct and are about the wrong place.
    That is the confidently-wrong failure this package exists to avoid, and it
    is invisible unless you check which portal each hit came from.
    """
    from urllib.parse import quote

    if not (domain or "").strip():
        raise NotConfigured(
            "discover_socrata() needs a portal domain: an empty one is not a "
            "narrower search, it is every Socrata portal in the country, and "
            "the hits would belong to other states.")
    url = (f"https://api.us.socrata.com/api/catalog/v1"
           f"?domains={quote(domain)}&q={quote(query)}&limit={int(limit)}")
    raw = _fetch(url, fetcher)
    try:
        blob = json.loads(raw)
    except ValueError as exc:
        raise SourceError(f"catalogue for {domain} did not return JSON: {exc}") from exc

    out: list[dict] = []
    for hit in (blob.get("results") or []):
        resource = hit.get("resource") or {}
        out.append({
            "id": resource.get("id") or "",
            "name": resource.get("name") or "",
            "updated": (resource.get("updatedAt") or "")[:10],
            "rows": resource.get("rows_count"),
            "link": hit.get("link") or "",
        })
    return [row for row in out if row["id"]]


# ---------------------------------------------------------------------------
# Delimited files and spreadsheets
# ---------------------------------------------------------------------------

def parse_delimited(data: bytes | str, *, delimiter: str = "") -> list[dict]:
    """CSV or TSV bytes into dicts keyed by header.

    Pure, so a state's real file can be dropped into a test verbatim. The
    delimiter is sniffed when not given, because states are inconsistent about
    it even between years of the same file.
    """
    text = data.decode("utf-8-sig", errors="replace") if isinstance(data, bytes) else data
    if not text.strip():
        return []

    if not delimiter:
        sample = text[:8192]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            # A single-column file sniffs as nothing; comma is the safe default.
            delimiter = "\t" if "\t" in sample.splitlines()[0] else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict] = []
    for row in reader:
        # Strip the header's stray whitespace and drop csv's None key for
        # ragged rows, which would otherwise blow up _pick's str() on the key.
        rows.append({(k or "").strip(): v
                     for k, v in row.items() if k is not None})
    return rows


def fetch_delimited(url: str, *, fetcher: Optional[Fetcher] = None,
                    delimiter: str = "") -> list[dict]:
    return parse_delimited(_fetch(url, fetcher), delimiter=delimiter)


def parse_xlsx(data: bytes, *, sheet: Optional[str] = None,
               header_row: int = 1) -> list[dict]:
    """First sheet of a workbook into dicts keyed by its header row.

    openpyxl is imported lazily and is not a hard dependency: a state that
    publishes CSV never pays for it, same reasoning as pypdf in campaigns_ca.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:      # pragma: no cover - environment-dependent
        raise SourceError(
            "openpyxl is required to read a .xlsx candidate file: "
            "pip install openpyxl") from exc

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = book[sheet] if sheet else book[book.sheetnames[0]]

    header: list[str] = []
    rows: list[dict] = []
    for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index < header_row:
            continue
        cells = ["" if v is None else str(v).strip() for v in values]
        if index == header_row:
            header = cells
            continue
        if not any(cells):
            continue
        rows.append(dict(zip(header, cells)))
    return rows


def fetch_xlsx(url: str, *, fetcher: Optional[Fetcher] = None,
               sheet: Optional[str] = None, header_row: int = 1) -> list[dict]:
    return parse_xlsx(_fetch(url, fetcher), sheet=sheet, header_row=header_row)


def fetch_table(url: str, *, fetcher: Optional[Fetcher] = None,
                **kw: Any) -> list[dict]:
    """Fetch a table, picking the reader from the URL's extension.

    A state that switches its download from .csv to .xlsx between cycles is
    common enough that dispatching on the URL costs less than two adapters.
    """
    lowered = url.lower().split("?")[0]
    if lowered.endswith((".xlsx", ".xlsm")):
        return fetch_xlsx(url, fetcher=fetcher,
                          sheet=kw.get("sheet"),
                          header_row=kw.get("header_row", 1))
    if lowered.endswith((".csv", ".tsv", ".txt")):
        return fetch_delimited(url, fetcher=fetcher,
                               delimiter=kw.get("delimiter", ""))
    # No usable extension: try delimited, which is what an unlabelled export
    # almost always is, and let SourceError carry the real shape on failure.
    return fetch_delimited(url, fetcher=fetcher, delimiter=kw.get("delimiter", ""))
